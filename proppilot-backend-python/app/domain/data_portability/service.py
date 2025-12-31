import io
import logging
from datetime import date
from decimal import Decimal
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.domain.data_portability.schemas import (
    ExcelImportPreview,
    ExcelImportResult,
    ImportAction,
    LeaseExcelRow,
    LeaseRowPreview,
    PaymentExcelRow,
    PaymentRowPreview,
    PropertyRowPreview,
    PropertyUnitExcelRow,
    TenantExcelRow,
    TenantRowPreview,
)
from app.domain.leases.models import AdjustmentIndex, Lease, LeaseStatus, lease_tenants
from app.domain.payments.models import Payment, PaymentStatus, PaymentType
from app.domain.properties.models import PropertyUnit
from app.domain.tenants.models import Tenant
from app.domain.users.models import User

logger = logging.getLogger(__name__)

# Sheet names in Spanish
SHEET_PROPERTIES = "Propiedades"
SHEET_TENANTS = "Inquilinos"
SHEET_LEASES = "Contratos"
SHEET_PAYMENTS = "Pagos"
SHEET_HELP = "Ayuda"

# Column headers
PROPERTY_HEADERS = [
    "Calle", "Numero", "Piso", "Depto", "Ciudad", "Provincia", "Codigo Postal", "Tipo", "Alquiler Base"
]
TENANT_HEADERS = ["Nombre Completo", "DNI/CUIT", "Email", "Telefono"]
LEASE_HEADERS = [
    "Direccion Propiedad", "DNI Inquilino", "Fecha Inicio", "Fecha Fin",
    "Alquiler Mensual", "Indice Ajuste", "Frecuencia Ajuste (meses)", "Estado"
]
PAYMENT_HEADERS = [
    "Direccion Propiedad", "DNI Inquilino", "Fecha Inicio Contrato",
    "Monto", "Fecha Pago", "Tipo", "Estado", "Descripcion"
]


class ExcelImportExportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_all(self, owner_id: int) -> bytes:
        """Export all user data to an Excel file with multiple sheets."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Export each entity type to its own sheet
        await self._export_properties(wb, owner_id)
        await self._export_tenants(wb, owner_id)
        await self._export_leases(wb, owner_id)
        await self._export_payments(wb, owner_id)

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_template(self) -> bytes:
        """Generate an empty Excel template with headers and help sheet."""
        wb = Workbook()
        wb.remove(wb.active)

        # Create empty sheets with headers
        self._create_sheet_with_headers(wb, SHEET_PROPERTIES, PROPERTY_HEADERS)
        self._create_sheet_with_headers(wb, SHEET_TENANTS, TENANT_HEADERS)
        self._create_sheet_with_headers(wb, SHEET_LEASES, LEASE_HEADERS)
        self._create_sheet_with_headers(wb, SHEET_PAYMENTS, PAYMENT_HEADERS)

        # Create help sheet
        self._create_help_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def preview_import(self, file: BinaryIO, owner_id: int) -> ExcelImportPreview:
        """Parse Excel file and return preview of what will be imported."""
        preview = ExcelImportPreview()

        wb = load_workbook(file, data_only=True)

        # Parse each sheet
        preview.properties = await self._parse_properties_sheet(wb, owner_id)
        preview.tenants = await self._parse_tenants_sheet(wb, owner_id)
        preview.leases = await self._parse_leases_sheet(
            wb, owner_id, preview.properties, preview.tenants
        )
        preview.payments = await self._parse_payments_sheet(wb, owner_id, preview.leases)

        preview.calculate_stats()
        return preview

    async def execute_import(
        self, preview: ExcelImportPreview, owner_id: int
    ) -> ExcelImportResult:
        """Execute the import after user confirms preview."""
        result = ExcelImportResult()

        # Get owner
        owner_result = await self.db.execute(select(User).where(User.id == owner_id))
        owner = owner_result.scalar_one_or_none()
        if not owner:
            result.error_messages.append("User not found")
            return result

        # Maps to track created entities for relationship linking
        created_properties: dict[str, PropertyUnit] = {}
        created_tenants: dict[str, Tenant] = {}
        created_leases: dict[str, Lease] = {}

        # Import properties first
        for row in preview.properties:
            if row.valid and row.action == ImportAction.CREATE:
                try:
                    property_unit = self._create_property_from_row(row.data, owner_id)
                    self.db.add(property_unit)
                    await self.db.flush()
                    created_properties[row.data.get_normalized_address()] = property_unit
                    result.properties_created += 1
                except Exception as e:
                    result.error_messages.append(f"Property row {row.row_number}: {e}")
                    result.errors += 1
            elif row.action == ImportAction.SKIP:
                result.skipped += 1

        # Import tenants
        for row in preview.tenants:
            if row.valid and row.action == ImportAction.CREATE:
                try:
                    tenant = self._create_tenant_from_row(row.data, owner_id)
                    self.db.add(tenant)
                    await self.db.flush()
                    created_tenants[row.data.national_id] = tenant
                    result.tenants_created += 1
                except Exception as e:
                    result.error_messages.append(f"Tenant row {row.row_number}: {e}")
                    result.errors += 1
            elif row.action == ImportAction.SKIP:
                result.skipped += 1

        # Import leases
        for row in preview.leases:
            if row.valid and row.action == ImportAction.CREATE:
                try:
                    lease = await self._create_lease_from_row(
                        row.data, owner_id, created_properties, created_tenants
                    )
                    self.db.add(lease)
                    await self.db.flush()
                    lease_key = self._build_lease_key(row.data)
                    created_leases[lease_key] = lease
                    result.leases_created += 1
                except Exception as e:
                    result.error_messages.append(f"Lease row {row.row_number}: {e}")
                    result.errors += 1
            elif row.action == ImportAction.SKIP:
                result.skipped += 1

        # Import payments
        for row in preview.payments:
            if row.valid and row.action == ImportAction.CREATE:
                try:
                    payment = await self._create_payment_from_row(
                        row.data, owner_id, created_leases
                    )
                    self.db.add(payment)
                    await self.db.flush()
                    result.payments_created += 1
                except Exception as e:
                    result.error_messages.append(f"Payment row {row.row_number}: {e}")
                    result.errors += 1
            elif row.action == ImportAction.SKIP:
                result.skipped += 1

        result.success = result.errors == 0
        return result

    # ==================== EXPORT METHODS ====================

    async def _export_properties(self, wb: Workbook, owner_id: int):
        ws = wb.create_sheet(SHEET_PROPERTIES)
        self._create_header_row(ws, PROPERTY_HEADERS)

        result = await self.db.execute(
            select(PropertyUnit).where(PropertyUnit.owner_id == owner_id)
        )
        properties = result.scalars().all()

        for i, p in enumerate(properties, start=2):
            ws.cell(row=i, column=1, value=p.street or "")
            ws.cell(row=i, column=2, value=p.street_number or "")
            ws.cell(row=i, column=3, value=p.floor or "")
            ws.cell(row=i, column=4, value=p.apartment or "")
            ws.cell(row=i, column=5, value=p.city or "")
            ws.cell(row=i, column=6, value=p.province or "")
            ws.cell(row=i, column=7, value=p.postal_code or "")
            ws.cell(row=i, column=8, value=p.type or "")
            ws.cell(row=i, column=9, value=float(p.base_rent_amount) if p.base_rent_amount else None)

        self._auto_size_columns(ws)

    async def _export_tenants(self, wb: Workbook, owner_id: int):
        ws = wb.create_sheet(SHEET_TENANTS)
        self._create_header_row(ws, TENANT_HEADERS)

        result = await self.db.execute(
            select(Tenant).where(Tenant.owner_id == owner_id)
        )
        tenants = result.scalars().all()

        for i, t in enumerate(tenants, start=2):
            ws.cell(row=i, column=1, value=t.full_name or "")
            ws.cell(row=i, column=2, value=t.national_id or "")
            ws.cell(row=i, column=3, value=t.email or "")
            ws.cell(row=i, column=4, value=t.phone or "")

        self._auto_size_columns(ws)

    async def _export_leases(self, wb: Workbook, owner_id: int):
        ws = wb.create_sheet(SHEET_LEASES)
        self._create_header_row(ws, LEASE_HEADERS)

        result = await self.db.execute(
            select(Lease)
            .options(
                selectinload(Lease.property_unit),
                selectinload(Lease.tenants),
            )
            .where(Lease.owner_id == owner_id, Lease.deleted == False)
        )
        leases = result.scalars().all()

        for i, l in enumerate(leases, start=2):
            ws.cell(row=i, column=1, value=l.property_unit.get_full_address() if l.property_unit else "")

            # Get tenant national IDs as comma-separated string
            tenant_ids = ", ".join(t.national_id for t in l.tenants)
            ws.cell(row=i, column=2, value=tenant_ids)

            ws.cell(row=i, column=3, value=l.start_date)
            ws.cell(row=i, column=4, value=l.end_date)
            ws.cell(row=i, column=5, value=float(l.monthly_rent) if l.monthly_rent else None)
            ws.cell(row=i, column=6, value=l.adjustment_index or "")
            ws.cell(row=i, column=7, value=l.adjustment_frequency_months or 12)
            ws.cell(row=i, column=8, value=l.status or "")

        self._auto_size_columns(ws)

    async def _export_payments(self, wb: Workbook, owner_id: int):
        ws = wb.create_sheet(SHEET_PAYMENTS)
        self._create_header_row(ws, PAYMENT_HEADERS)

        result = await self.db.execute(
            select(Payment)
            .options(
                selectinload(Payment.lease).selectinload(Lease.property_unit),
                selectinload(Payment.lease).selectinload(Lease.tenants),
            )
            .where(Payment.owner_id == owner_id)
        )
        payments = result.scalars().all()

        for i, p in enumerate(payments, start=2):
            lease = p.lease
            ws.cell(
                row=i, column=1,
                value=lease.property_unit.get_full_address() if lease and lease.property_unit else ""
            )

            tenant_id = ""
            if lease and lease.tenants:
                first_tenant = next(iter(lease.tenants), None)
                tenant_id = first_tenant.national_id if first_tenant else ""
            ws.cell(row=i, column=2, value=tenant_id)

            ws.cell(row=i, column=3, value=lease.start_date if lease else None)
            ws.cell(row=i, column=4, value=float(p.amount) if p.amount else None)
            ws.cell(row=i, column=5, value=p.payment_date)
            ws.cell(row=i, column=6, value=p.payment_type or "")
            ws.cell(row=i, column=7, value=p.status or "")
            ws.cell(row=i, column=8, value=p.description or "")

        self._auto_size_columns(ws)

    # ==================== PARSE METHODS ====================

    async def _parse_properties_sheet(
        self, wb: Workbook, owner_id: int
    ) -> list[PropertyRowPreview]:
        result = []
        if SHEET_PROPERTIES not in wb.sheetnames:
            return result

        ws = wb[SHEET_PROPERTIES]

        # Load existing properties for duplicate detection
        existing_result = await self.db.execute(
            select(PropertyUnit).where(PropertyUnit.owner_id == owner_id)
        )
        existing_properties = existing_result.scalars().all()
        existing_addresses = {p.address.lower().strip() for p in existing_properties if p.address}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            data = PropertyUnitExcelRow(
                street=self._get_string_value(row[0]) if len(row) > 0 else None,
                streetNumber=self._get_string_value(row[1]) if len(row) > 1 else None,
                floor=self._get_string_value(row[2]) if len(row) > 2 else None,
                apartment=self._get_string_value(row[3]) if len(row) > 3 else None,
                city=self._get_string_value(row[4]) if len(row) > 4 else None,
                province=self._get_string_value(row[5]) if len(row) > 5 else None,
                postalCode=self._get_string_value(row[6]) if len(row) > 6 else None,
                type=self._get_string_value(row[7]) if len(row) > 7 else None,
                baseRentAmount=self._get_decimal_value(row[8]) if len(row) > 8 else None,
            )

            preview = PropertyRowPreview(rowNumber=row_num, data=data, errors=[], warnings=[])

            # Validate
            self._validate_property_row(data, preview)

            # Check for duplicates
            if not preview.errors:
                normalized_address = data.get_normalized_address()
                if normalized_address in existing_addresses:
                    preview.action = ImportAction.SKIP
                    preview.warnings.append("Ya existe una propiedad con esta direccion")
                else:
                    preview.action = ImportAction.CREATE
                    existing_addresses.add(normalized_address)
                preview.valid = True
            else:
                preview.action = ImportAction.ERROR
                preview.valid = False

            result.append(preview)

        return result

    async def _parse_tenants_sheet(
        self, wb: Workbook, owner_id: int
    ) -> list[TenantRowPreview]:
        result = []
        if SHEET_TENANTS not in wb.sheetnames:
            return result

        ws = wb[SHEET_TENANTS]

        existing_result = await self.db.execute(
            select(Tenant).where(Tenant.owner_id == owner_id)
        )
        existing_tenants = existing_result.scalars().all()
        existing_national_ids = {t.national_id for t in existing_tenants}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            data = TenantExcelRow(
                fullName=self._get_string_value(row[0]) if len(row) > 0 else None,
                nationalId=self._get_string_value(row[1]) if len(row) > 1 else None,
                email=self._get_string_value(row[2]) if len(row) > 2 else None,
                phone=self._get_string_value(row[3]) if len(row) > 3 else None,
            )

            preview = TenantRowPreview(rowNumber=row_num, data=data, errors=[], warnings=[])

            self._validate_tenant_row(data, preview)

            if not preview.errors:
                if data.national_id in existing_national_ids:
                    preview.action = ImportAction.SKIP
                    preview.warnings.append("Ya existe un inquilino con este DNI/CUIT")
                else:
                    preview.action = ImportAction.CREATE
                    existing_national_ids.add(data.national_id)
                preview.valid = True
            else:
                preview.action = ImportAction.ERROR
                preview.valid = False

            result.append(preview)

        return result

    async def _parse_leases_sheet(
        self,
        wb: Workbook,
        owner_id: int,
        property_previews: list[PropertyRowPreview],
        tenant_previews: list[TenantRowPreview],
    ) -> list[LeaseRowPreview]:
        result = []
        if SHEET_LEASES not in wb.sheetnames:
            return result

        ws = wb[SHEET_LEASES]

        # Build lookup sets for validation
        available_properties = set()
        existing_props = await self.db.execute(
            select(PropertyUnit).where(PropertyUnit.owner_id == owner_id)
        )
        for p in existing_props.scalars().all():
            if p.address:
                available_properties.add(p.address.lower().strip())
        for pp in property_previews:
            if pp.valid and pp.action == ImportAction.CREATE:
                available_properties.add(pp.data.get_normalized_address())

        available_tenants = set()
        existing_tenants = await self.db.execute(
            select(Tenant).where(Tenant.owner_id == owner_id)
        )
        for t in existing_tenants.scalars().all():
            available_tenants.add(t.national_id)
        for tp in tenant_previews:
            if tp.valid and tp.action == ImportAction.CREATE:
                available_tenants.add(tp.data.national_id)

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            data = LeaseExcelRow(
                propertyAddress=self._get_string_value(row[0]) if len(row) > 0 else None,
                tenantNationalIds=self._get_string_value(row[1]) if len(row) > 1 else None,
                startDate=self._get_date_value(row[2]) if len(row) > 2 else None,
                endDate=self._get_date_value(row[3]) if len(row) > 3 else None,
                monthlyRent=self._get_decimal_value(row[4]) if len(row) > 4 else None,
                adjustmentIndex=self._get_string_value(row[5]) if len(row) > 5 else None,
                adjustmentFrequencyMonths=self._get_int_value(row[6]) if len(row) > 6 else None,
                status=self._get_string_value(row[7]) if len(row) > 7 else None,
            )

            preview = LeaseRowPreview(rowNumber=row_num, data=data, errors=[], warnings=[])

            self._validate_lease_row(data, preview, available_properties, available_tenants)

            if not preview.errors:
                preview.action = ImportAction.CREATE
                preview.valid = True
            else:
                preview.action = ImportAction.ERROR
                preview.valid = False

            result.append(preview)

        return result

    async def _parse_payments_sheet(
        self,
        wb: Workbook,
        owner_id: int,
        lease_previews: list[LeaseRowPreview],
    ) -> list[PaymentRowPreview]:
        result = []
        if SHEET_PAYMENTS not in wb.sheetnames:
            return result

        ws = wb[SHEET_PAYMENTS]

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            data = PaymentExcelRow(
                propertyAddress=self._get_string_value(row[0]) if len(row) > 0 else None,
                tenantNationalId=self._get_string_value(row[1]) if len(row) > 1 else None,
                leaseStartDate=self._get_date_value(row[2]) if len(row) > 2 else None,
                amount=self._get_decimal_value(row[3]) if len(row) > 3 else None,
                paymentDate=self._get_date_value(row[4]) if len(row) > 4 else None,
                paymentType=self._get_string_value(row[5]) if len(row) > 5 else None,
                status=self._get_string_value(row[6]) if len(row) > 6 else None,
                description=self._get_string_value(row[7]) if len(row) > 7 else None,
            )

            preview = PaymentRowPreview(rowNumber=row_num, data=data, errors=[], warnings=[])

            self._validate_payment_row(data, preview)

            if not preview.errors:
                preview.action = ImportAction.CREATE
                preview.valid = True
            else:
                preview.action = ImportAction.ERROR
                preview.valid = False

            result.append(preview)

        return result

    # ==================== VALIDATION METHODS ====================

    def _validate_property_row(self, data: PropertyUnitExcelRow, preview: PropertyRowPreview):
        if not data.street:
            preview.errors.append("Calle es requerida")
        if not data.street_number:
            preview.errors.append("Numero es requerido")
        if not data.type:
            preview.errors.append("Tipo es requerido")
        if not data.base_rent_amount or data.base_rent_amount <= 0:
            preview.errors.append("Alquiler base debe ser mayor a 0")

    def _validate_tenant_row(self, data: TenantExcelRow, preview: TenantRowPreview):
        if not data.full_name:
            preview.errors.append("Nombre completo es requerido")
        if not data.national_id:
            preview.errors.append("DNI/CUIT es requerido")
        if not data.email:
            preview.errors.append("Email es requerido")
        if not data.phone:
            preview.errors.append("Telefono es requerido")

    def _validate_lease_row(
        self,
        data: LeaseExcelRow,
        preview: LeaseRowPreview,
        available_properties: set,
        available_tenants: set,
    ):
        if not data.property_address:
            preview.errors.append("Direccion propiedad es requerida")
        else:
            normalized = data.property_address.lower().strip()
            if normalized not in available_properties:
                preview.errors.append(f"Propiedad no encontrada: {data.property_address}")

        if not data.tenant_national_ids:
            preview.errors.append("DNI inquilino es requerido")
        else:
            for dni in data.tenant_national_ids.split(","):
                if dni.strip() not in available_tenants:
                    preview.errors.append(f"Inquilino no encontrado: {dni.strip()}")

        if not data.start_date:
            preview.errors.append("Fecha inicio es requerida")
        if not data.end_date:
            preview.errors.append("Fecha fin es requerida")
        if data.start_date and data.end_date and data.end_date <= data.start_date:
            preview.errors.append("Fecha fin debe ser posterior a fecha inicio")
        if not data.monthly_rent or data.monthly_rent <= 0:
            preview.errors.append("Alquiler mensual debe ser mayor a 0")

        # Validate adjustment index
        if data.adjustment_index:
            try:
                AdjustmentIndex(data.adjustment_index.upper())
            except ValueError:
                preview.errors.append(f"Indice de ajuste invalido: {data.adjustment_index}")

        # Validate status
        if data.status:
            try:
                LeaseStatus(data.status.upper())
            except ValueError:
                preview.errors.append(f"Estado invalido: {data.status}")

    def _validate_payment_row(self, data: PaymentExcelRow, preview: PaymentRowPreview):
        if not data.property_address:
            preview.errors.append("Direccion propiedad es requerida")
        if not data.tenant_national_id:
            preview.errors.append("DNI inquilino es requerido")
        if not data.lease_start_date:
            preview.errors.append("Fecha inicio contrato es requerida")
        if not data.amount or data.amount <= 0:
            preview.errors.append("Monto debe ser mayor a 0")
        if not data.payment_date:
            preview.errors.append("Fecha pago es requerida")

        # Validate payment type
        if data.payment_type:
            try:
                PaymentType(data.payment_type.upper())
            except ValueError:
                preview.errors.append(f"Tipo de pago invalido: {data.payment_type}")

        # Validate status
        if data.status:
            try:
                PaymentStatus(data.status.upper())
            except ValueError:
                preview.errors.append(f"Estado invalido: {data.status}")

    # ==================== ENTITY CREATION METHODS ====================

    def _create_property_from_row(self, data: PropertyUnitExcelRow, owner_id: int) -> PropertyUnit:
        return PropertyUnit(
            street=data.street,
            street_number=data.street_number,
            floor=data.floor,
            apartment=data.apartment,
            city=data.city,
            province=data.province,
            postal_code=data.postal_code,
            type=data.type,
            base_rent_amount=data.base_rent_amount,
            address=data.get_full_address(),
            owner_id=owner_id,
        )

    def _create_tenant_from_row(self, data: TenantExcelRow, owner_id: int) -> Tenant:
        return Tenant(
            full_name=data.full_name,
            national_id=data.national_id,
            email=data.email,
            phone=data.phone,
            owner_id=owner_id,
        )

    async def _create_lease_from_row(
        self,
        data: LeaseExcelRow,
        owner_id: int,
        created_properties: dict[str, PropertyUnit],
        created_tenants: dict[str, Tenant],
    ) -> Lease:
        # Find property
        normalized_address = data.property_address.lower().strip()
        property_unit = created_properties.get(normalized_address)

        if not property_unit:
            result = await self.db.execute(
                select(PropertyUnit).where(
                    PropertyUnit.owner_id == owner_id,
                    PropertyUnit.address.ilike(f"%{data.property_address}%"),
                )
            )
            property_unit = result.scalars().first()

        if not property_unit:
            raise ValueError(f"Propiedad no encontrada: {data.property_address}")

        # Find tenants
        tenants = []
        for dni in data.tenant_national_ids.split(","):
            trimmed_dni = dni.strip()
            tenant = created_tenants.get(trimmed_dni)
            if not tenant:
                result = await self.db.execute(
                    select(Tenant).where(
                        Tenant.national_id == trimmed_dni,
                        Tenant.owner_id == owner_id,
                    )
                )
                tenant = result.scalars().first()
            if not tenant:
                raise ValueError(f"Inquilino no encontrado: {trimmed_dni}")
            tenants.append(tenant)

        lease = Lease(
            property_unit_id=property_unit.id,
            start_date=data.start_date,
            end_date=data.end_date,
            monthly_rent=data.monthly_rent,
            owner_id=owner_id,
        )

        if data.adjustment_index:
            lease.adjustment_index = AdjustmentIndex(data.adjustment_index.upper())
        if data.adjustment_frequency_months:
            lease.adjustment_frequency_months = data.adjustment_frequency_months
        if data.status:
            lease.status = LeaseStatus(data.status.upper())

        lease.tenants = tenants
        return lease

    async def _create_payment_from_row(
        self,
        data: PaymentExcelRow,
        owner_id: int,
        created_leases: dict[str, Lease],
    ) -> Payment:
        # Find lease
        lease_key = f"{data.property_address.lower().strip()}|{data.tenant_national_id}|{data.lease_start_date}"
        lease = created_leases.get(lease_key)

        if not lease:
            # Try to find existing lease
            result = await self.db.execute(
                select(Lease)
                .options(
                    selectinload(Lease.property_unit),
                    selectinload(Lease.tenants),
                )
                .where(Lease.owner_id == owner_id, Lease.deleted == False)
            )
            leases = result.scalars().all()

            for l in leases:
                if (
                    l.property_unit
                    and data.property_address.lower() in l.property_unit.address.lower()
                    and any(t.national_id == data.tenant_national_id for t in l.tenants)
                    and l.start_date == data.lease_start_date
                ):
                    lease = l
                    break

        if not lease:
            raise ValueError("Contrato no encontrado para el pago")

        payment = Payment(
            lease_id=lease.id,
            amount=data.amount,
            payment_date=data.payment_date,
            description=data.description,
            owner_id=owner_id,
        )

        if data.payment_type:
            payment.payment_type = PaymentType(data.payment_type.upper())
        if data.status:
            payment.status = PaymentStatus(data.status.upper())

        return payment

    def _build_lease_key(self, data: LeaseExcelRow) -> str:
        first_tenant = data.tenant_national_ids.split(",")[0].strip()
        return f"{data.property_address.lower().strip()}|{first_tenant}|{data.start_date}"

    # ==================== HELPER METHODS ====================

    def _create_sheet_with_headers(self, wb: Workbook, sheet_name: str, headers: list[str]):
        ws = wb.create_sheet(sheet_name)
        self._create_header_row(ws, headers)
        self._auto_size_columns(ws)

    def _create_header_row(self, ws, headers: list[str]):
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

    def _create_help_sheet(self, wb: Workbook):
        ws = wb.create_sheet(SHEET_HELP)

        help_text = [
            ("INSTRUCCIONES PARA IMPORTAR DATOS", True),
            ("", False),
            ("1. Complete cada hoja con los datos correspondientes", False),
            ("2. Respete el formato de fechas: AAAA-MM-DD (ejemplo: 2024-01-15)", False),
            ("3. Los montos deben ser numeros positivos", False),
            ("", False),
            ("VALORES VALIDOS:", True),
            ("", False),
            ("Tipo de propiedad: apartment, house, duplex, ph, studio, loft", False),
            ("Indice de ajuste: ICL, IPC, DOLAR_BLUE, DOLAR_OFICIAL, DOLAR_MEP, NONE", False),
            ("Estado contrato: ACTIVE, EXPIRED, TERMINATED", False),
            ("Tipo de pago: RENT, DEPOSIT, MAINTENANCE, UTILITY, OTHER", False),
            ("Estado pago: PAID, PENDING", False),
        ]

        for row_num, (text, bold) in enumerate(help_text, start=1):
            cell = ws.cell(row=row_num, column=1, value=text)
            if bold:
                cell.font = Font(bold=True)

        ws.column_dimensions["A"].width = 80

    def _auto_size_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _get_string_value(self, cell) -> str | None:
        if cell is None:
            return None
        if isinstance(cell, str):
            return cell.strip() if cell.strip() else None
        return str(cell).strip() if cell else None

    def _get_decimal_value(self, cell) -> Decimal | None:
        if cell is None:
            return None
        try:
            if isinstance(cell, (int, float)):
                return Decimal(str(cell))
            if isinstance(cell, str):
                return Decimal(cell.strip().replace(",", "."))
        except Exception:
            pass
        return None

    def _get_int_value(self, cell) -> int | None:
        if cell is None:
            return None
        try:
            if isinstance(cell, (int, float)):
                return int(cell)
            if isinstance(cell, str):
                return int(cell.strip())
        except Exception:
            pass
        return None

    def _get_date_value(self, cell) -> date | None:
        if cell is None:
            return None
        try:
            if isinstance(cell, date):
                return cell
            if hasattr(cell, "date"):
                return cell.date()
            if isinstance(cell, str):
                return date.fromisoformat(cell.strip())
        except Exception:
            pass
        return None
